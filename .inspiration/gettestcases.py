​import requests
import sys
import os
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from base64 import b64encode
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json

# Config 
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))) 
import PAT 
from GetAzDoToken.GetAzCliToken import get_azure_cli_token

plan_id = 10519879
organization = PAT.ORGANIZATION 
project = PAT.PROJECT 
token = b64encode(f':{get_azure_cli_token()}'.encode()).decode()

headers = { 
    "Authorization": f"Basic {token}",
    "Content-Type": "application/json"
}

num_workers = 100

# Debug-Modus: Wenn True, werden rohe JSON-Daten exportiert
DEBUG_MODE = False


def safe_get_display_name(field_value):
    """Extrahiert displayName aus einem Feld."""
    if isinstance(field_value, dict):
        return field_value.get("displayName", "")
    return str(field_value) if field_value else ""


def safe_get_unique_name(field_value):
    """Extrahiert uniqueName aus einem Feld."""
    if isinstance(field_value, dict):
        return field_value.get("uniqueName", "")
    return ""


def safe_get_id(field_value):
    """Extrahiert id aus einem Feld."""
    if isinstance(field_value, dict):
        return field_value.get("id", "")
    return ""


def get_all_suites(plan_id): 
    """Ruft alle Test Suites eines Plans als Baumstruktur ab."""
    url = f"https://dev.azure.com/{organization}/{project}/_apis/test/Plans/{plan_id}/suites?$asTreeView=true&api-version=5.0" 
    response = requests.get(url, headers=headers) 
    if response.status_code == 200: 
        data = response.json() 
        return data.get("value", []) 
    else: 
        print(f"Fehler beim Abrufen der Test Suites: {response.status_code} - {response.text}") 
        return [] 


def iterate_suites(suites, parent_path=""): 
    """Durchlaeuft die Liste der Suites rekursiv."""
    result = [] 
    for suite in suites: 
        current_path = f"{parent_path} > {suite['name']}" if parent_path else suite["name"] 
        result.append((suite, current_path)) 
        if "children" in suite and suite["children"]: 
            result.extend(iterate_suites(suite["children"], current_path)) 
    return result 


def get_test_cases_for_suite(plan_id, suite_id):
    """Ruft die Test Cases fuer eine einzelne Suite ab."""
    url = f"https://dev.azure.com/{organization}/{project}/_apis/test/Plans/{plan_id}/suites/{suite_id}/testcases?api-version=5.0"
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        data = response.json()
        return data.get("value", [])
    else:
        print(f"Fehler beim Abrufen der Test Cases fuer Suite {suite_id}: {response.status_code} - {response.text}")
        return []


def get_test_points_for_suite(plan_id, suite_id):
    """Ruft alle Test Points fuer eine Suite ab (mit Paging)."""
    all_points = []
    continuation_token = None

    while True:
        url = f"https://dev.azure.com/{organization}/{project}/_apis/test/Plans/{plan_id}/suites/{suite_id}/points?includePointDetails=true&api-version=7.1"
        if continuation_token:
            url += f"&continuationToken={continuation_token}"

        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            data = response.json()
            points = data.get("value", [])

            if not points:
                break

            all_points.extend(points)

            # Pruefe auf Continuation Token im Header
            continuation_token = response.headers.get("x-ms-continuationtoken")
            if not continuation_token:
                break
        else:
            print(f"Fehler beim Abrufen der Test Points fuer Suite {suite_id}: {response.status_code} - {response.text}")
            break

    return all_points


def fetch_points_for_suite(suite_tuple):
    """Worker-Funktion fuer parallelen Abruf von Test Points."""
    suite, suite_path = suite_tuple
    suite_id = suite["id"]
    suite_name = suite["name"]
    points = get_test_points_for_suite(plan_id, suite_id)
    # Fuege Suite-Info zu jedem Point hinzu
    for point in points:
        point["_suite_id"] = suite_id
        point["_suite_name"] = suite_name
        point["_suite_path"] = suite_path
    return points


def get_all_test_points_parallel(suite_list):
    """Ruft Test Points fuer alle Suites parallel ab."""
    all_points = []
    total_suites = len(suite_list)

    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        futures = {executor.submit(fetch_points_for_suite, suite_tuple): suite_tuple for suite_tuple in suite_list}
        completed = 0

        for future in as_completed(futures):
            completed += 1
            suite_tuple = futures[future]
            suite, _ = suite_tuple
            suite_name = suite["name"]

            try:
                points = future.result()
                all_points.extend(points)
                if completed % 10 == 0 or completed == total_suites:
                    print(f"  Points abgerufen: {completed}/{total_suites} Suites, {len(all_points)} Points gesamt")
            except Exception as e:
                print(f"  Fehler bei Suite {suite_name}: {e}")

    return all_points


def extract_point_fields(point):
    """Extrahiert alle relevanten Felder aus einem Test Point."""

    test_case = point.get("testCase", {}) or {}
    configuration = point.get("configuration", {}) or {}
    last_test_run = point.get("lastTestRun", {}) or {}
    last_result = point.get("lastResult", {}) or {}
    assigned_to = point.get("assignedTo", {}) or {}
    last_updated_by = point.get("lastUpdatedBy", {}) or {}

    row = {
        # Basis Point Felder
        "PointID": point.get("id", ""),
        "URL": point.get("url", ""),

        # Test Case Referenz
        "TestCase_ID": test_case.get("id", ""),
        "TestCase_Name": test_case.get("name", ""),
        "TestCase_URL": test_case.get("url", ""),

        # Konfiguration
        "Configuration_ID": configuration.get("id", ""),
        "Configuration_Name": configuration.get("name", ""),

        # Status
        "State": point.get("state", ""),
        "Outcome": point.get("outcome", ""),

        # Letzter Test Run
        "LastTestRun_ID": last_test_run.get("id", ""),
        "LastTestRun_Name": last_test_run.get("name", ""),

        # Letztes Result
        "LastResult_ID": last_result.get("id", ""),
        "LastResult_State": last_result.get("state", ""),

        # Benutzer: Assigned To
        "AssignedTo_DisplayName": safe_get_display_name(assigned_to),
        "AssignedTo_UniqueName": safe_get_unique_name(assigned_to),
        "AssignedTo_ID": safe_get_id(assigned_to),

        # Benutzer: Last Updated By
        "LastUpdatedBy_DisplayName": safe_get_display_name(last_updated_by),
        "LastUpdatedBy_UniqueName": safe_get_unique_name(last_updated_by),
        "LastUpdatedBy_ID": safe_get_id(last_updated_by),

        # Zeitangaben
        "LastUpdatedDate": point.get("lastUpdatedDate", ""),
        "LastResetToActive": point.get("lastResetToActive", ""),
        "LastRunBuildNumber": point.get("lastRunBuildNumber", ""),

        # Suite Info (aus _suite_* Feldern)
        "SuiteID": point.get("_suite_id", ""),
        "Suite": point.get("_suite_name", ""),
        "SuitePath": point.get("_suite_path", ""),

        # Weitere Felder
        "Comment": point.get("comment", ""),
        "IsAutomated": point.get("isAutomated", ""),
    }

    return row 


def fetch_chunk(chunk):
    """Ruft fuer einen Chunk von Work Item IDs die Details ab."""
    chunk_details = {}
    ids_str = ','.join(map(str, chunk))
    url = f"https://dev.azure.com/{organization}/{project}/_apis/wit/workitems?ids={ids_str}&api-version=7.0"
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        data = response.json()
        for item in data.get("value", []):
            item_id = item.get("id")
            chunk_details[int(item_id)] = item
            chunk_details[str(item_id)] = item
    else:
        print(f"Fehler beim Abrufen von Work Items: {response.status_code} - {response.text}")
    
    return chunk_details


def get_work_items_details_batch(work_item_ids):
    """Ruft die Details mehrerer Work Items ab."""
    if not work_item_ids:
        return {}
    
    details = {}
    chunk_size = 200
    chunks = [work_item_ids[i:i+chunk_size] for i in range(0, len(work_item_ids), chunk_size)]
    
    with ThreadPoolExecutor() as executor:
        futures = {executor.submit(fetch_chunk, chunk): chunk for chunk in chunks}
        for future in as_completed(futures):
            details.update(future.result())
    
    return details


def get_all_runs_for_plan(plan_id):
    """Ruft alle Test Runs fuer einen Plan ab (mit Paging)."""
    all_runs = []
    skip = 0
    top = 1000  # Max pro Request

    while True:
        url = f"https://dev.azure.com/{organization}/{project}/_apis/test/runs?planId={plan_id}&$top={top}&$skip={skip}&api-version=7.1"
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            data = response.json()
            runs = data.get("value", [])

            if not runs:
                break

            all_runs.extend(runs)
            print(f"  Runs abgerufen: {len(all_runs)} (skip={skip})")

            if len(runs) < top:
                break

            skip += top
        else:
            print(f"Fehler beim Abrufen der Runs: {response.status_code} - {response.text}")
            break

    return all_runs


def get_results_for_run(run_id):
    """Ruft alle Test Results fuer einen Run ab (mit Paging)."""
    all_results = []
    skip = 0
    top = 100  # Max pro Request mit detailsToInclude (max 1000 ohne)

    while True:
        url = f"https://dev.azure.com/{organization}/{project}/_apis/test/Runs/{run_id}/results?$top={top}&$skip={skip}&detailsToInclude=Point&api-version=7.1"
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            data = response.json()
            results = data.get("value", [])

            if not results:
                break

            all_results.extend(results)

            if len(results) < top:
                break

            skip += top
        else:
            print(f"Fehler beim Abrufen der Results fuer Run {run_id}: {response.status_code} - {response.text}")
            break

    return all_results


def fetch_results_for_run(run):
    """Worker-Funktion fuer parallelen Abruf von Test Results."""
    run_id = run.get("id")
    results = get_results_for_run(run_id)
    return results


def get_all_results_for_runs(runs):
    """Ruft Test Results fuer alle Runs parallel ab."""
    all_results = []
    total_runs = len(runs)

    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        futures = {executor.submit(fetch_results_for_run, run): run for run in runs}
        completed = 0

        for future in as_completed(futures):
            completed += 1
            run = futures[future]
            run_id = run.get("id")

            try:
                results = future.result()
                all_results.extend(results)
                if completed % 50 == 0 or completed == total_runs:
                    print(f"  Results abgerufen: {completed}/{total_runs} Runs verarbeitet, {len(all_results)} Results gesamt")
            except Exception as e:
                print(f"  Fehler bei Run {run_id}: {e}")

    return all_results


def extract_result_fields(result):
    """Extrahiert alle relevanten Felder aus einem Test Result."""

    test_run = result.get("testRun", {}) or {}
    test_case = result.get("testCase", {}) or {}
    test_plan = result.get("testPlan", {}) or {}
    test_suite = result.get("testSuite", {}) or {}
    test_point = result.get("testPoint", {}) or {}
    configuration = result.get("configuration", {}) or {}
    build = result.get("build", {}) or {}
    build_ref = result.get("buildReference", {}) or {}
    release = result.get("release", {}) or {}
    release_ref = result.get("releaseReference", {}) or {}
    run_by = result.get("runBy", {}) or {}
    owner = result.get("owner", {}) or {}
    last_updated_by = result.get("lastUpdatedBy", {}) or {}
    area = result.get("area", {}) or {}
    failing_since = result.get("failingSince", {}) or {}
    project_ref = result.get("project", {}) or {}

    row = {
        # === WICHTIGE MATCHING-FELDER ===
        # Diese Felder werden zum Verknuepfen mit anderen Tabellen verwendet
        "TestCaseReferenceId": result.get("testCaseReferenceId", ""),  # = WorkItemID in Test Cases Tabelle
        "RunID": test_run.get("id", ""),  # = RunID in Test Runs Tabelle
        "TestSuiteID": test_suite.get("id", ""),
        "TestPlanID": test_plan.get("id", ""),
        "TestPointID": test_point.get("id", ""),

        # Basis Result Felder
        "ResultID": result.get("id", ""),
        "URL": result.get("url", ""),

        # Test Case Referenz
        "TestCase_ID": test_case.get("id", ""),
        "TestCase_Name": test_case.get("name", ""),
        "TestCaseTitle": result.get("testCaseTitle", ""),
        "TestCaseRevision": result.get("testCaseRevision", ""),

        # Test Run Referenz
        "TestRun_Name": test_run.get("name", ""),
        "TestRun_URL": test_run.get("url", ""),

        # Test Plan/Suite Referenz
        "TestPlan_Name": test_plan.get("name", ""),
        "TestSuite_Name": test_suite.get("name", ""),
        "TestPoint_Name": test_point.get("name", ""),

        # Status und Ergebnis
        "State": result.get("state", ""),
        "Outcome": result.get("outcome", ""),
        "Priority": result.get("priority", ""),

        # Zeitangaben
        "StartedDate": result.get("startedDate", ""),
        "CompletedDate": result.get("completedDate", ""),
        "DurationInMs": result.get("durationInMs", ""),
        "CreatedDate": result.get("createdDate", ""),
        "LastUpdatedDate": result.get("lastUpdatedDate", ""),

        # Konfiguration
        "Configuration_ID": configuration.get("id", ""),
        "Configuration_Name": configuration.get("name", ""),
        "ComputerName": result.get("computerName", ""),

        # Build Referenz
        "Build_ID": build.get("id", ""),
        "Build_Name": build.get("name", ""),
        "Build_URL": build.get("url", ""),
        "BuildRef_ID": build_ref.get("id", ""),
        "BuildRef_Number": build_ref.get("number", ""),
        "BuildRef_DefinitionId": build_ref.get("definitionId", ""),
        "BuildRef_BranchName": build_ref.get("branchName", ""),

        # Release Referenz
        "Release_ID": release.get("id", ""),
        "Release_Name": release.get("name", ""),
        "Release_URL": release.get("url", ""),
        "ReleaseRef_ID": release_ref.get("id", ""),
        "ReleaseRef_Name": release_ref.get("name", ""),
        "ReleaseRef_EnvironmentName": release_ref.get("environmentName", ""),

        # Fehler-Informationen
        "ErrorMessage": result.get("errorMessage", ""),
        "StackTrace": result.get("stackTrace", ""),
        "FailureType": result.get("failureType", ""),
        "ResolutionState": result.get("resolutionState", ""),
        "ResolutionStateId": result.get("resolutionStateId", ""),

        # Failing Since
        "FailingSince_Date": failing_since.get("date", ""),

        # Automation Felder
        "AutomatedTestId": result.get("automatedTestId", ""),
        "AutomatedTestName": result.get("automatedTestName", ""),
        "AutomatedTestStorage": result.get("automatedTestStorage", ""),
        "AutomatedTestType": result.get("automatedTestType", ""),
        "AutomatedTestTypeId": result.get("automatedTestTypeId", ""),

        # Benutzer: Run By
        "RunBy_DisplayName": safe_get_display_name(run_by),
        "RunBy_UniqueName": safe_get_unique_name(run_by),
        "RunBy_ID": safe_get_id(run_by),

        # Benutzer: Owner
        "Owner_DisplayName": safe_get_display_name(owner),
        "Owner_UniqueName": safe_get_unique_name(owner),
        "Owner_ID": safe_get_id(owner),

        # Benutzer: Last Updated By
        "LastUpdatedBy_DisplayName": safe_get_display_name(last_updated_by),
        "LastUpdatedBy_UniqueName": safe_get_unique_name(last_updated_by),
        "LastUpdatedBy_ID": safe_get_id(last_updated_by),

        # Area
        "Area_ID": area.get("id", ""),
        "Area_Name": area.get("name", ""),

        # Projekt
        "Project_ID": project_ref.get("id", ""),
        "Project_Name": project_ref.get("name", ""),

        # Sonstige
        "Comment": result.get("comment", ""),
        "Revision": result.get("revision", ""),
    }

    # Associated Bugs als kommaseparierte Liste
    associated_bugs = result.get("associatedBugs", []) or []
    bug_ids = [str(bug.get("id", "")) for bug in associated_bugs if bug.get("id")]
    row["AssociatedBugIDs"] = ", ".join(bug_ids)

    return row


def extract_run_fields(run):
    """Extrahiert alle relevanten Felder aus einem Test Run."""
    
    build_config = run.get("buildConfiguration", {}) or {}
    plan = run.get("plan", {}) or {}
    owner = run.get("owner", {}) or {}
    project_ref = run.get("project", {}) or {}
    
    row = {
        # Basis Run Felder
        "RunID": run.get("id", ""),
        "Name": run.get("name", ""),
        "URL": run.get("url", ""),
        "WebAccessUrl": run.get("webAccessUrl", ""),
        
        # Status und Ergebnis
        "State": run.get("state", ""),
        "IsAutomated": run.get("isAutomated", ""),
        
        # Statistiken
        "TotalTests": run.get("totalTests", ""),
        "PassedTests": run.get("passedTests", ""),
        "IncompleteTests": run.get("incompleteTests", ""),
        "NotApplicableTests": run.get("notApplicableTests", ""),
        "UnanalyzedTests": run.get("unanalyzedTests", ""),
        
        # Zeitangaben
        "StartedDate": run.get("startedDate", ""),
        "CompletedDate": run.get("completedDate", ""),
        "CreatedDate": run.get("createdDate", ""),
        "LastUpdatedDate": run.get("lastUpdatedDate", ""),
        
        # Build Konfiguration
        "BuildConfiguration_ID": build_config.get("id", ""),
        "BuildConfiguration_Number": build_config.get("number", ""),
        "BuildConfiguration_Flavor": build_config.get("flavor", ""),
        "BuildConfiguration_Platform": build_config.get("platform", ""),
        "BuildConfiguration_URI": build_config.get("uri", ""),
        
        # Plan Referenz
        "Plan_ID": plan.get("id", ""),
        "Plan_Name": plan.get("name", ""),
        
        # Owner
        "Owner_DisplayName": safe_get_display_name(owner),
        "Owner_UniqueName": safe_get_unique_name(owner),
        "Owner_ID": safe_get_id(owner),
        
        # Projekt
        "Project_ID": project_ref.get("id", ""),
        "Project_Name": project_ref.get("name", ""),
        
        # Weitere Felder
        "Revision": run.get("revision", ""),
        "RunType": run.get("runType", ""),
        "PostProcessState": run.get("postProcessState", ""),
        "DtlAutEnvironment": run.get("dtlAutEnvironment", {}).get("id", "") if run.get("dtlAutEnvironment") else "",
        "Comment": run.get("comment", ""),
    }
    
    return row


def extract_all_fields(wi, tc, suite_id, suite_name, suite_path):
    """Extrahiert alle relevanten Felder aus dem Work Item und Test Case."""

    if not wi or not isinstance(wi, dict):
        return None

    fields = wi.get("fields", {})
    links = wi.get("_links", {})

    row = {
        # Basis Work Item Felder
        "WorkItemID": wi.get("id", ""),
        "Rev": wi.get("rev", ""),
        "URL": wi.get("url", ""),

        # System Felder
        "AreaPath": fields.get("System.AreaPath", ""),
        "TeamProject": fields.get("System.TeamProject", ""),
        "IterationPath": fields.get("System.IterationPath", ""),
        "WorkItemType": fields.get("System.WorkItemType", ""),
        "State": fields.get("System.State", ""),
        "Reason": fields.get("System.Reason", ""),
        "Title": fields.get("System.Title", ""),
        "CommentCount": fields.get("System.CommentCount", ""),
        "Tags": fields.get("System.Tags", ""),

        # AssignedTo Felder
        "AssignedTo_DisplayName": safe_get_display_name(fields.get("System.AssignedTo")),
        "AssignedTo_UniqueName": safe_get_unique_name(fields.get("System.AssignedTo")),
        "AssignedTo_ID": safe_get_id(fields.get("System.AssignedTo")),

        # CreatedBy Felder
        "CreatedDate": fields.get("System.CreatedDate", ""),
        "CreatedBy_DisplayName": safe_get_display_name(fields.get("System.CreatedBy")),
        "CreatedBy_UniqueName": safe_get_unique_name(fields.get("System.CreatedBy")),
        "CreatedBy_ID": safe_get_id(fields.get("System.CreatedBy")),

        # ChangedBy Felder
        "ChangedDate": fields.get("System.ChangedDate", ""),
        "ChangedBy_DisplayName": safe_get_display_name(fields.get("System.ChangedBy")),
        "ChangedBy_UniqueName": safe_get_unique_name(fields.get("System.ChangedBy")),
        "ChangedBy_ID": safe_get_id(fields.get("System.ChangedBy")),

        # Microsoft VSTS Common Felder
        "StateChangeDate": fields.get("Microsoft.VSTS.Common.StateChangeDate", ""),
        "ActivatedDate": fields.get("Microsoft.VSTS.Common.ActivatedDate", ""),
        "ActivatedBy_DisplayName": safe_get_display_name(fields.get("Microsoft.VSTS.Common.ActivatedBy")),
        "ActivatedBy_UniqueName": safe_get_unique_name(fields.get("Microsoft.VSTS.Common.ActivatedBy")),
        "ActivatedBy_ID": safe_get_id(fields.get("Microsoft.VSTS.Common.ActivatedBy")),
        "Priority": fields.get("Microsoft.VSTS.Common.Priority", ""),

        # TCM Felder (Test Case Management)
        "AutomationStatus": fields.get("Microsoft.VSTS.TCM.AutomationStatus", ""),
        "Steps": fields.get("Microsoft.VSTS.TCM.Steps", ""),

        # Links
        "Link_Self": links.get("self", {}).get("href", "") if isinstance(links.get("self"), dict) else "",
        "Link_HTML": links.get("html", {}).get("href", "") if isinstance(links.get("html"), dict) else "",
        "Link_WorkItemUpdates": links.get("workItemUpdates", {}).get("href", "") if isinstance(links.get("workItemUpdates"), dict) else "",
        "Link_WorkItemRevisions": links.get("workItemRevisions", {}).get("href", "") if isinstance(links.get("workItemRevisions"), dict) else "",
        "Link_WorkItemComments": links.get("workItemComments", {}).get("href", "") if isinstance(links.get("workItemComments"), dict) else "",
        "Link_WorkItemType": links.get("workItemType", {}).get("href", "") if isinstance(links.get("workItemType"), dict) else "",
        "Link_Fields": links.get("fields", {}).get("href", "") if isinstance(links.get("fields"), dict) else "",

        # Suite Informationen
        "SuiteID": suite_id,
        "Suite": suite_name,
        "SuitePath": suite_path,

        # Test Case spezifische Felder
        "TestCaseURL": tc.get("testCase", {}).get("url", ""),
    }

    return row


def process_suite(suite, suite_path):
    """Verarbeitet eine einzelne Suite."""
    suite_id = suite["id"]
    suite_name = suite["name"]

    test_cases = get_test_cases_for_suite(plan_id, suite_id)

    if not test_cases:
        return []

    # Sammle Work Item IDs - als int
    work_item_ids = []
    tc_map = {}

    for tc in test_cases:
        test_case_ref = tc.get("testCase", {})
        work_item_id = test_case_ref.get("id")
        if work_item_id:
            work_item_id_int = int(work_item_id)
            work_item_ids.append(work_item_id_int)
            tc_map[work_item_id_int] = tc

    if not work_item_ids:
        return []

    # Batch-Abruf der Work Items
    wi_details = get_work_items_details_batch(work_item_ids)

    local_rows = []
    for work_item_id in work_item_ids:
        wi = wi_details.get(work_item_id) or wi_details.get(str(work_item_id))
        tc = tc_map.get(work_item_id)

        if not wi:
            continue

        row = extract_all_fields(wi, tc, suite_id, suite_name, suite_path)
        if row:
            local_rows.append(row)

    return local_rows


def process_all_suites_parallel(suite_list):
    """Verarbeitet alle Suites parallel."""
    all_rows = []
    total_suites = len(suite_list)

    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        # Submit alle Suite-Verarbeitungen
        futures = {
            executor.submit(process_suite, suite, suite_path): (idx, suite["name"])
            for idx, (suite, suite_path) in enumerate(suite_list, start=1)
        }

        completed = 0
        for future in as_completed(futures):
            completed += 1
            idx, suite_name = futures[future]

            try:
                rows = future.result()
                all_rows.extend(rows)
                if completed % 10 == 0 or completed == total_suites:
                    print(f"  Suites verarbeitet: {completed}/{total_suites}, {len(all_rows)} Test Cases gesamt")
            except Exception as e:
                print(f"  Fehler bei Suite {suite_name}: {e}")

    return all_rows


def parse_date(date_str):
    """Parst ein Datum aus verschiedenen Formaten und gibt ein date-Objekt zurueck."""
    if not date_str:
        return None
    try:
        # Azure DevOps Format: 2024-01-15T10:30:00.000Z oder 2024-01-15T10:30:00Z
        if isinstance(date_str, str):
            # Entferne Millisekunden und Z
            clean = date_str.split('.')[0].replace('Z', '')
            dt = datetime.strptime(clean, "%Y-%m-%dT%H:%M:%S")
            return dt.date()
    except (ValueError, AttributeError):
        pass
    return None


def parse_datetime(date_str):
    """Parst ein Datum aus verschiedenen Formaten und gibt ein datetime-Objekt zurueck (inkl. Uhrzeit)."""
    if not date_str:
        return None
    try:
        # Azure DevOps Format: 2024-01-15T10:30:00.000Z oder 2024-01-15T10:30:00Z
        if isinstance(date_str, str):
            # Entferne Millisekunden und Z
            clean = date_str.split('.')[0].replace('Z', '')
            dt = datetime.strptime(clean, "%Y-%m-%dT%H:%M:%S")
            return dt
    except (ValueError, AttributeError):
        pass
    return None


def generate_daily_snapshots(testcase_rows, result_rows):
    """
    Generiert eine Snapshot-Tabelle mit dem Status jedes Testfalls fuer jeden Tag.
    Beruecksichtigt, dass ein Test Case in mehreren Suites sein kann.

    Ausgabe: Liste von Dicts mit:
    - Datum
    - TestCaseID (WorkItemID)
    - SuiteID
    - Suite (Name)
    - Title
    - Status (Outcome des letzten Results bis zu diesem Datum, oder 'NotRun')
    - TestResultID (ID des relevanten Results, oder leer)
    - RunID
    """
    print("\n===== Generiere Daily Snapshots =====")

    if not testcase_rows:
        print("Keine Test Cases vorhanden.")
        return []

    # Erstelle Mapping: (WorkItemID, SuiteID) -> Info
    # Ein Test Case kann in mehreren Suites sein!
    testcase_info = {}
    for tc in testcase_rows:
        tc_id = tc.get("WorkItemID")
        suite_id = tc.get("SuiteID")
        if not tc_id or not suite_id:
            continue
        created = parse_date(tc.get("CreatedDate"))
        title = tc.get("Title", "")
        suite_name = tc.get("Suite", "")
        if created:
            key = (int(tc_id), int(suite_id))
            testcase_info[key] = {
                "created": created,
                "title": title,
                "suite_name": suite_name
            }

    print(f"  {len(testcase_info)} Test Case/Suite Kombinationen mit gueltigem CreatedDate")

    # Erstelle Mapping: (TestCase_ID, TestSuiteID) -> Liste von Results
    # Matching: TestCase_ID = WorkItemID, TestSuiteID = SuiteID
    results_by_testcase_suite = {}

    for res in result_rows:
        tc_id_raw = res.get("TestCase_ID")
        suite_id_raw = res.get("TestSuiteID")
        if not tc_id_raw or not suite_id_raw:
            continue

        key = (int(tc_id_raw), int(suite_id_raw))
        completed_date = parse_date(res.get("CompletedDate"))
        completed_datetime = parse_datetime(res.get("CompletedDate"))
        outcome = res.get("Outcome", "")
        result_id = res.get("ResultID", "")
        run_id = res.get("RunID", "")

        if completed_date and completed_datetime and outcome:
            if key not in results_by_testcase_suite:
                results_by_testcase_suite[key] = []
            results_by_testcase_suite[key].append({
                "date": completed_date,
                "datetime": completed_datetime,  # Vollstaendiges datetime fuer korrekte Sortierung
                "outcome": outcome,
                "result_id": result_id,
                "run_id": run_id
            })

    # Sortiere Results pro (TestCase, Suite) nach vollstaendigem datetime (inkl. Uhrzeit)
    for key in results_by_testcase_suite:
        results_by_testcase_suite[key].sort(key=lambda x: x["datetime"])

    # Statistik
    testcase_keys = set(testcase_info.keys())
    result_keys = set(results_by_testcase_suite.keys())
    matching_keys = testcase_keys & result_keys
    print(f"  {len(results_by_testcase_suite)} unique (TestCase, Suite) Kombinationen in Results")
    print(f"  {len(matching_keys)} Kombinationen mit Results gematched")

    # Bestimme Datumsbereich
    all_dates = []
    for key, info in testcase_info.items():
        all_dates.append(info["created"])
    for key, results in results_by_testcase_suite.items():
        for r in results:
            all_dates.append(r["date"])

    if not all_dates:
        print("Keine Datumswerte gefunden.")
        return []

    min_date = min(all_dates)
    max_date = datetime.now().date()

    print(f"  Datumsbereich: {min_date} bis {max_date}")

    # Worker-Funktion fuer parallele Snapshot-Generierung
    def generate_snapshots_for_testcase_suite(key, info, tc_results, max_date):
        """Generiert Snapshots fuer einen einzelnen Testcase in einer Suite."""
        tc_id, suite_id = key
        rows = []
        created_date = info["created"]
        title = info["title"]
        suite_name = info["suite_name"]

        current_date = created_date
        current_status = "NotRun"
        current_result_id = ""
        current_run_id = ""
        result_idx = 0

        while current_date <= max_date:
            while result_idx < len(tc_results) and tc_results[result_idx]["date"] <= current_date:
                current_status = tc_results[result_idx]["outcome"]
                current_result_id = tc_results[result_idx]["result_id"]
                current_run_id = tc_results[result_idx]["run_id"]
                result_idx += 1

            rows.append({
                "Datum": current_date.strftime("%Y-%m-%d"),
                "TestCaseID": tc_id,
                "SuiteID": suite_id,
                "Suite": suite_name,
                "Title": title,
                "Status": current_status,
                "TestResultID": current_result_id,
                "RunID": current_run_id
            })

            current_date += timedelta(days=1)

        return rows

    # Generiere Snapshots parallel
    snapshot_rows = []
    total_combinations = len(testcase_info)

    print(f"  Generiere Snapshots fuer {total_combinations} (TestCase, Suite) Kombinationen parallel...")

    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        futures = {
            executor.submit(
                generate_snapshots_for_testcase_suite,
                key,
                info,
                results_by_testcase_suite.get(key, []),
                max_date
            ): key
            for key, info in testcase_info.items()
        }

        completed = 0
        for future in as_completed(futures):
            completed += 1
            try:
                rows = future.result()
                snapshot_rows.extend(rows)
                if completed % 100 == 0 or completed == total_combinations:
                    print(f"  Verarbeitet: {completed}/{total_combinations} Kombinationen, {len(snapshot_rows)} Zeilen")
            except Exception as e:
                key = futures[future]
                print(f"  Fehler bei {key}: {e}")

    print(f"  Gesamt: {len(snapshot_rows)} Snapshot-Zeilen generiert")
    return snapshot_rows


def write_debug_json(data, output_file):
    """Schreibt rohe Daten als JSON-Datei fuer Debug-Zwecke."""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    print(f"DEBUG: JSON-Datei '{output_file}' mit {len(data)} Eintraegen wurde erstellt.")


def write_to_xlsx(all_rows, output_file, sheet_name="Data"):
    """Schreibt die Daten in eine XLSX-Datei."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if not all_rows:
        print(f"Keine Daten zum Schreiben vorhanden fuer {output_file}.")
        wb.save(output_file)
        return
    
    fieldnames = list(all_rows[0].keys())
    
    # Header schreiben
    for col_idx, field in enumerate(fieldnames, start=1):
        ws.cell(row=1, column=col_idx, value=field)
    
    # Daten schreiben
    for row_idx, row_data in enumerate(all_rows, start=2):
        for col_idx, field in enumerate(fieldnames, start=1):
            value = row_data.get(field, "")
            if isinstance(value, str) and len(value) > 32767:
                value = value[:32764] + "..."
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Spaltenbreiten anpassen
    for col_idx, field in enumerate(fieldnames, start=1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = min(max(len(field) + 2, 12), 50)
    
    wb.save(output_file)
    print(f"XLSX-Datei '{output_file}' mit {len(all_rows)} Eintraegen wurde erstellt.")


def main(): 
    print(f"Organisation: {organization}")
    print(f"Projekt: {project}")
    print(f"Plan ID: {plan_id}")
    print("-" * 50)
    
    base_path = rf"{PAT.USERPATH}\ING\Projects CoE I&A - Dokumente\General\DE Datacenter Lifecycle - Modernization SCS-, Windows and JEE Platforms\Projekt Lifecycle Linux & JEE Platform Upgrade\BitBucket-Jenkins Decom\Rohdaten"
    
    # ===== TEIL 1: Test Cases =====
    print("\n===== TEIL 1: Test Cases abrufen =====")
    suites = get_all_suites(plan_id)
    if not suites:
        print("Keine Test Suites gefunden.")
    else:
        suite_list = iterate_suites(suites)
        total_suites = len(suite_list)
        print(f"Gefunden: {total_suites} Suites")
        print(f"Verarbeite parallel mit {num_workers} Workers...")
        print("-" * 50)

        all_testcase_rows = process_all_suites_parallel(suite_list)

        # DEBUG: Rohe Suites und Test Cases als JSON exportieren
        if DEBUG_MODE:
            debug_file_suites = rf"{base_path}\debug_raw_suites.json"
            write_debug_json(suites, debug_file_suites)

        output_file_testcases = rf"{base_path}\testplan_testcases.xlsx"
        write_to_xlsx(all_testcase_rows, output_file_testcases, "Test Cases")

        # ===== TEIL 1b: Test Points =====
        print("\n===== TEIL 1b: Test Points abrufen =====")
        print(f"Rufe Points fuer {total_suites} Suites ab (parallel mit {num_workers} Workers)...")
        all_points = get_all_test_points_parallel(suite_list)
        print(f"Gesamt: {len(all_points)} Points gefunden")

        if all_points:
            output_file_points = rf"{base_path}\testplan_points.json"
            write_debug_json(all_points, output_file_points)

    # ===== TEIL 2: Test Runs =====
    print("\n===== TEIL 2: Test Runs abrufen =====")
    runs = get_all_runs_for_plan(plan_id)
    print(f"Gesamt: {len(runs)} Runs gefunden")

    if runs:
        # DEBUG: Rohe Runs als JSON exportieren
        if DEBUG_MODE:
            debug_file_runs = rf"{base_path}\debug_raw_runs.json"
            write_debug_json(runs, debug_file_runs)

        all_run_rows = []
        for run in runs:
            row = extract_run_fields(run)
            all_run_rows.append(row)

        output_file_runs = rf"{base_path}\testplan_runs.xlsx"
        write_to_xlsx(all_run_rows, output_file_runs, "Test Runs")

        # ===== TEIL 3: Test Results =====
        print("\n===== TEIL 3: Test Results abrufen =====")
        print(f"Rufe Results fuer {len(runs)} Runs ab (parallel mit {num_workers} Workers)...")
        all_results = get_all_results_for_runs(runs)
        print(f"Gesamt: {len(all_results)} Results gefunden")

        if all_results:
            # DEBUG: Rohe Results als JSON exportieren
            if DEBUG_MODE:
                debug_file_results = rf"{base_path}\debug_raw_results.json"
                write_debug_json(all_results, debug_file_results)

                # Zusätzlich: Nur Results wo testSuite leer ist
                results_without_suite = [r for r in all_results if not r.get("testSuite") or not r.get("testSuite", {}).get("id")]
                if results_without_suite:
                    debug_file_no_suite = rf"{base_path}\debug_results_without_suite.json"
                    write_debug_json(results_without_suite, debug_file_no_suite)
                    print(f"DEBUG: {len(results_without_suite)} Results haben keine testSuite ID!")

            all_result_rows = []
            for result in all_results:
                row = extract_result_fields(result)
                all_result_rows.append(row)

            output_file_results = rf"{base_path}\testplan_results.xlsx"
            write_to_xlsx(all_result_rows, output_file_results, "Test Results")

            # ===== TEIL 4: Daily Snapshots =====
            print("\n===== TEIL 4: Daily Snapshots generieren =====")
            snapshot_rows = generate_daily_snapshots(all_testcase_rows, all_result_rows)

            if snapshot_rows:
                output_file_snapshots = rf"{base_path}\testplan_daily_snapshots.xlsx"
                write_to_xlsx(snapshot_rows, output_file_snapshots, "Daily Snapshots")

            # Info zum Matching ausgeben
            print("\n" + "-" * 50)
            print("MATCHING-INFO:")
            print("  Test Results <-> Test Cases:")
            print("    TestResults.TestCase_ID = TestCases.WorkItemID")
            print("    TestResults.TestSuiteID = TestCases.SuiteID")
            print("  Test Results <-> Test Runs:")
            print("    TestResults.RunID = TestRuns.RunID")
            print("  Test Points <-> Test Cases:")
            print("    TestPoints.TestCase_ID = TestCases.WorkItemID")
            print("    TestPoints.SuiteID = TestCases.SuiteID")
            print("  Test Points <-> Test Results:")
            print("    TestPoints.PointID = TestResults.TestPointID")
            print("  Daily Snapshots:")
            print("    Snapshots.TestCaseID = TestCases.WorkItemID")
            print("    Snapshots.SuiteID = TestCases.SuiteID")
            print("    Snapshots.TestResultID = TestResults.ResultID")
            print("    Snapshots.RunID = TestResults.RunID")
        else:
            print("Keine Results gefunden.")
    else:
        print("Keine Runs gefunden.")

    print("\n" + "=" * 50)
    print("Fertig!")
    input("Druecke Enter, um das Programm zu beenden...")


if __name__ == "__main__":
    main()